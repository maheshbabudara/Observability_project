import openpyxl as p

def reader(filename,sheet_name):
    workbook=p.load_workbook(filename)
    sheet=workbook[sheet_name]
    rows=sheet.max_row
    cols=sheet.max_column
    details=[]
    for row in range(2,rows+1):
        data=[]
        for col in range(1,cols+1):
            data.append(sheet.cell(row,col).value)
        details.append(data)
    return details
